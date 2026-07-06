"""
Builds the KPI workbook (Excel) that accompanies the case study report.
Uses openpyxl; rates/percentages that can be derived from raw counts in the
same sheet are written as formulas (per xlsx skill conventions), not
hardcoded values.
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

DATA = "/sessions/serene-amazing-cori/work/data"
OUT = "/sessions/serene-amazing-cori/work/dashboard/Regional_Fulfillment_KPI_Workbook.xlsx"

NAVY = "2B2620"      # charcoal-ink (primary dark, no blue)
TEAL = "4B6B43"      # forest/olive green (secondary)
WHITE = "FFFFFF"
LIGHT = "F2F2F2"
BLUE_INPUT = "A0522D"  # sienna/rust — used for hardcoded input cells instead of blue

FONT = "Arial"
header_font = Font(name=FONT, bold=True, color=WHITE, size=11)
header_fill = PatternFill("solid", start_color=NAVY, end_color=NAVY)
title_font = Font(name=FONT, bold=True, color=NAVY, size=14)
subtitle_font = Font(name=FONT, italic=True, color="555555", size=10)
normal_font = Font(name=FONT, size=10)
input_font = Font(name=FONT, size=10, color=BLUE_INPUT)
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()
wb.remove(wb.active)


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def autosize(ws, ncols, widths=None):
    for c in range(1, ncols + 1):
        letter = get_column_letter(c)
        ws.column_dimensions[letter].width = widths[c - 1] if widths else 20


def write_df(ws, df, start_row=1, pct_cols=None, money_cols=None, int_cols=None, widths=None):
    pct_cols = pct_cols or []
    money_cols = money_cols or []
    int_cols = int_cols or []
    ncols = len(df.columns)
    for j, col in enumerate(df.columns, start=1):
        ws.cell(row=start_row, column=j, value=col)
    style_header(ws, start_row, ncols)
    for i, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
        for j, col in enumerate(df.columns, start=1):
            val = row[col]
            cell = ws.cell(row=i, column=j, value=val)
            cell.font = normal_font
            cell.border = border
            if col in pct_cols:
                cell.number_format = "0.0%"
            elif col in money_cols:
                cell.number_format = "$#,##0;($#,##0);-"
            elif col in int_cols:
                cell.number_format = "#,##0;(#,##0);-"
            if (i - start_row) % 2 == 0:
                cell.fill = PatternFill("solid", start_color=LIGHT, end_color=LIGHT)
    autosize(ws, ncols, widths)
    return start_row + len(df) + 1


def title_block(ws, title, subtitle):
    ws["A1"] = title
    ws["A1"].font = title_font
    ws["A2"] = subtitle
    ws["A2"].font = subtitle_font
    return 4


# =====================================================================
# SHEET 0: Read Me
# =====================================================================
ws = wb.create_sheet("Read Me")
ws.column_dimensions["A"].width = 110
rows = [
    ("Regional Outbound Fulfillment & Logistics Analytics — KPI Workbook", title_font),
    ("Companion data workbook to the case study report and interactive dashboard.", subtitle_font),
    ("", normal_font),
    ("DATA SOURCE", Font(name=FONT, bold=True, size=11, color=NAVY)),
    ("DataCo Smart Supply Chain Dataset (Kaggle). kagglehub id: shashwatwork/dataco-smart-supply-chain-for-big-data-analysis", normal_font),
    ("178,626 orders analyzed (of 180,519 raw records; fraud-hold/payment-review rows that never shipped were excluded), Jan 2015 - Jan 2018.", normal_font),
    ("", normal_font),
    ("METHODOLOGY NOTE", Font(name=FONT, bold=True, size=11, color=NAVY)),
    ("Cost per Order, margin leakage, and savings-simulation figures are a transparent, documented parcel-cost model layered over real", normal_font),
    ("order-level operational fields (shipping mode, scheduled vs. actual transit days, region, order value, quantity). No public dataset", normal_font),
    ("exposes real carrier invoices, so this model stands in for one, exactly as an analyst would build a cost estimate ahead of finalized", normal_font),
    ("carrier rate cards. Full formulas are documented in the case study report Appendix A.", normal_font),
    ("", normal_font),
    ("HOW TO READ THIS WORKBOOK", Font(name=FONT, bold=True, size=11, color=NAVY)),
    ("Blue text = input values carried over from the Python analysis (counts, rates as measured in the data).", normal_font),
    ("Black text = formulas computed live within this workbook from the blue inputs (e.g., rates recomputed from counts, cumulative %, totals).", normal_font),
    ("", normal_font),
    ("SHEET INDEX", Font(name=FONT, bold=True, size=11, color=NAVY)),
    ("1. Topline KPIs — headline metrics for the whole network", normal_font),
    ("2. OTIF by Carrier — On-Time-In-Full performance by service tier", normal_font),
    ("3. OTIF by Region — OTIF performance by order region", normal_font),
    ("4. Cost per Order — modeled shipping cost by region x carrier", normal_font),
    ("5. EDD Accuracy — Estimated Delivery Date performance by carrier", normal_font),
    ("6. Failed Delivery Pareto — root cause by region and category", normal_font),
    ("7. Cross-Border USCA — Domestic US vs. US to Canada performance", normal_font),
    ("8. Corporate Segment — B2B seasonal volume and category mix", normal_font),
    ("9. Savings Simulation — cost-reduction levers and simulated impact", normal_font),
]
for i, (text, font) in enumerate(rows, start=1):
    ws.cell(row=i, column=1, value=text).font = font

# =====================================================================
# SHEET 1: Topline KPIs (with live formulas)
# =====================================================================
ws = wb.create_sheet("Topline KPIs")
r = title_block(ws, "Topline KPIs", "Headline fulfillment & cost metrics across all 178,626 analyzed orders")
summary = pd.read_csv(f"{DATA}/kpi_topline_summary.csv", index_col=0).iloc[:, 0]

labels = [
    ("Total Orders Analyzed", int(float(summary["total_orders"])), "#,##0"),
    ("OTIF Orders (derived)", round(float(summary["otif_rate"]) * float(summary["total_orders"])), "#,##0"),
    ("On-Time Orders (derived)", round(float(summary["on_time_rate"]) * float(summary["total_orders"])), "#,##0"),
    ("Cancelled Orders (derived)", round(float(summary["cancel_rate"]) * float(summary["total_orders"])), "#,##0"),
    ("Total Outbound Shipping Cost ($)", float(summary["total_shipping_cost"]), "$#,##0"),
    ("Total Shipping Margin Leakage ($)", float(summary["total_margin_leakage"]), "$#,##0"),
    ("Total Reshipment Cost ($, cancelled orders)", float(summary["total_reshipment_cost"]), "$#,##0"),
    ("Simulated Annual Savings ($, 3 levers)", float(summary["simulated_annual_savings"]), "$#,##0"),
    ("Corporate (B2B) Segment Orders", int(float(summary["corporate_segment_orders"])), "#,##0"),
    ("Corporate (B2B) Segment Sales ($)", float(summary["corporate_segment_sales"]), "$#,##0"),
    ("Cross-Border US->Canada Orders", int(float(summary["usca_cross_border_orders"])), "#,##0"),
]
row = r
ws.cell(row=row, column=1, value="Metric").font = header_font
ws.cell(row=row, column=1).fill = header_fill
ws.cell(row=row, column=2, value="Value").font = header_font
ws.cell(row=row, column=2).fill = header_fill
row += 1
input_start = row
for label, val, fmt in labels:
    ws.cell(row=row, column=1, value=label).font = normal_font
    c = ws.cell(row=row, column=2, value=val)
    c.font = input_font
    c.number_format = fmt
    row += 1

# derived formula rows
row += 1
ws.cell(row=row, column=1, value="Derived Rates (computed by formula from rows above)").font = Font(name=FONT, bold=True, size=11, color=NAVY)
row += 1
derived = [
    ("OTIF Rate", f"=B{input_start+1}/B{input_start}", "0.0%"),
    ("On-Time Rate", f"=B{input_start+2}/B{input_start}", "0.0%"),
    ("Cancellation Rate", f"=B{input_start+3}/B{input_start}", "0.0%"),
    ("Avg. Cost per Order", f"=B{input_start+4}/B{input_start}", "$#,##0.00"),
    ("Margin Leakage as % of Total Cost", f"=B{input_start+5}/B{input_start+4}", "0.0%"),
    ("Simulated Savings as % of Total Cost", f"=B{input_start+7}/B{input_start+4}", "0.0%"),
]
for label, formula, fmt in derived:
    ws.cell(row=row, column=1, value=label).font = normal_font
    c = ws.cell(row=row, column=2, value=formula)
    c.font = normal_font
    c.number_format = fmt
    row += 1
autosize(ws, 2, [42, 22])

# =====================================================================
# SHEET 2: OTIF by Carrier (formulas)
# =====================================================================
ws = wb.create_sheet("OTIF by Carrier")
r = title_block(ws, "OTIF by Carrier / Service Tier", "OTIF rate recomputed by formula from order counts")
df = pd.read_csv(f"{DATA}/kpi_otif_by_carrier.csv")
df["otif_orders"] = (df["otif_rate"] * df["orders"]).round().astype(int)
df["on_time_orders"] = (df["on_time_rate"] * df["orders"]).round().astype(int)
df["cancelled_orders"] = (df["cancel_rate"] * df["orders"]).round().astype(int)
out = df[["Shipping Mode", "orders", "otif_orders", "on_time_orders", "cancelled_orders", "avg_transit_variance", "avg_cost_per_order"]].copy()
out.columns = ["Carrier / Service Tier", "Orders", "OTIF Orders", "On-Time Orders", "Cancelled Orders", "Avg Transit Variance (days)", "Avg Cost/Order ($)"]
hdr_row = r
for j, col in enumerate(out.columns, start=1):
    ws.cell(row=hdr_row, column=j, value=col)
style_header(ws, hdr_row, len(out.columns))
start_data = hdr_row + 1
for i, (_, row_) in enumerate(out.iterrows()):
    rr = start_data + i
    ws.cell(row=rr, column=1, value=row_["Carrier / Service Tier"]).font = normal_font
    c = ws.cell(row=rr, column=2, value=int(row_["Orders"])); c.font = input_font; c.number_format = "#,##0"
    c = ws.cell(row=rr, column=3, value=int(row_["OTIF Orders"])); c.font = input_font; c.number_format = "#,##0"
    c = ws.cell(row=rr, column=4, value=int(row_["On-Time Orders"])); c.font = input_font; c.number_format = "#,##0"
    c = ws.cell(row=rr, column=5, value=int(row_["Cancelled Orders"])); c.font = input_font; c.number_format = "#,##0"
    c = ws.cell(row=rr, column=6, value=round(float(row_["Avg Transit Variance (days)"]), 2)); c.font = input_font; c.number_format = "0.00"
    c = ws.cell(row=rr, column=7, value=round(float(row_["Avg Cost/Order ($)"]), 2)); c.font = input_font; c.number_format = "$#,##0.00"
    for c_ in range(1, 8):
        ws.cell(row=rr, column=c_).border = border
        if i % 2 == 1:
            ws.cell(row=rr, column=c_).fill = PatternFill("solid", start_color=LIGHT, end_color=LIGHT)
end_data = start_data + len(out) - 1
# formula columns for rates
ws.cell(row=hdr_row, column=8, value="OTIF Rate")
ws.cell(row=hdr_row, column=9, value="On-Time Rate")
ws.cell(row=hdr_row, column=10, value="Cancel Rate")
style_header(ws, hdr_row, 10)
for i in range(start_data, end_data + 1):
    ws.cell(row=i, column=8, value=f"=C{i}/B{i}").number_format = "0.0%"
    ws.cell(row=i, column=9, value=f"=D{i}/B{i}").number_format = "0.0%"
    ws.cell(row=i, column=10, value=f"=E{i}/B{i}").number_format = "0.0%"
    for c_ in (8, 9, 10):
        ws.cell(row=i, column=c_).font = normal_font
        ws.cell(row=i, column=c_).border = border
total_row = end_data + 1
ws.cell(row=total_row, column=1, value="TOTAL / NETWORK AVG").font = Font(name=FONT, bold=True)
ws.cell(row=total_row, column=2, value=f"=SUM(B{start_data}:B{end_data})").font = Font(name=FONT, bold=True)
ws.cell(row=total_row, column=2).number_format = "#,##0"
ws.cell(row=total_row, column=3, value=f"=SUM(C{start_data}:C{end_data})").font = Font(name=FONT, bold=True)
ws.cell(row=total_row, column=3).number_format = "#,##0"
ws.cell(row=total_row, column=8, value=f"=C{total_row}/B{total_row}").font = Font(name=FONT, bold=True)
ws.cell(row=total_row, column=8).number_format = "0.0%"
autosize(ws, 10, [20, 10, 12, 14, 14, 20, 16, 12, 14, 12])

# =====================================================================
# SHEET 3: OTIF by Region
# =====================================================================
ws = wb.create_sheet("OTIF by Region")
r = title_block(ws, "OTIF by Order Region", "")
df = pd.read_csv(f"{DATA}/kpi_otif_by_region.csv")
df = df.rename(columns={"Order Region": "Order Region", "orders": "Orders", "otif_rate": "OTIF Rate",
                         "on_time_rate": "On-Time Rate", "cancel_rate": "Cancel Rate", "avg_cost_per_order": "Avg Cost/Order ($)"})
write_df(ws, df, start_row=r, pct_cols=["OTIF Rate", "On-Time Rate", "Cancel Rate"], money_cols=["Avg Cost/Order ($)"],
         int_cols=["Orders"], widths=[20, 10, 12, 12, 12, 16])

# =====================================================================
# SHEET 4: Cost per Order (region x carrier)
# =====================================================================
ws = wb.create_sheet("Cost per Order")
r = title_block(ws, "Cost per Order — by Region x Carrier", "Modeled outbound shipping cost (see report Appendix A for formula)")
df = pd.read_csv(f"{DATA}/kpi_cpo_region_carrier.csv")
df = df.rename(columns={"Order Region": "Order Region", "Shipping Mode": "Carrier", "orders": "Orders",
                         "avg_cost_per_order": "Avg Cost/Order ($)", "total_cost": "Total Cost ($)",
                         "avg_margin_leakage": "Avg Margin Leakage ($)", "total_margin_leakage": "Total Margin Leakage ($)"})
write_df(ws, df, start_row=r, money_cols=["Avg Cost/Order ($)", "Total Cost ($)", "Avg Margin Leakage ($)", "Total Margin Leakage ($)"],
         int_cols=["Orders"], widths=[20, 16, 10, 16, 16, 18, 20])

# =====================================================================
# SHEET 5: EDD Accuracy
# =====================================================================
ws = wb.create_sheet("EDD Accuracy")
r = title_block(ws, "EDD Accuracy by Carrier", "EDD-met rate recomputed by formula from order counts")
df = pd.read_csv(f"{DATA}/kpi_edd_by_carrier.csv")
df["edd_met_orders"] = (df["edd_met_rate"] * df["orders"]).round().astype(int)
out = df[["Shipping Mode", "orders", "edd_met_orders", "avg_scheduled_days", "avg_real_days", "avg_variance"]].copy()
out.columns = ["Carrier", "Orders", "EDD Met Orders", "Avg Scheduled Days", "Avg Real Days", "Avg Variance (days)"]
hdr_row = r
for j, col in enumerate(out.columns, start=1):
    ws.cell(row=hdr_row, column=j, value=col)
style_header(ws, hdr_row, len(out.columns) + 1)
start_data = hdr_row + 1
for i, (_, row_) in enumerate(out.iterrows()):
    rr = start_data + i
    ws.cell(row=rr, column=1, value=row_["Carrier"]).font = normal_font
    for j, col in enumerate(["Orders", "EDD Met Orders"], start=2):
        c = ws.cell(row=rr, column=j, value=int(row_[col])); c.font = input_font; c.number_format = "#,##0"
    for j, col in enumerate(["Avg Scheduled Days", "Avg Real Days", "Avg Variance (days)"], start=4):
        c = ws.cell(row=rr, column=j, value=round(float(row_[col]), 2)); c.font = input_font; c.number_format = "0.00"
    for c_ in range(1, 7):
        ws.cell(row=rr, column=c_).border = border
        if i % 2 == 1:
            ws.cell(row=rr, column=c_).fill = PatternFill("solid", start_color=LIGHT, end_color=LIGHT)
end_data = start_data + len(out) - 1
ws.cell(row=hdr_row, column=7, value="EDD Met Rate")
style_header(ws, hdr_row, 7)
for i in range(start_data, end_data + 1):
    c = ws.cell(row=i, column=7, value=f"=C{i}/B{i}")
    c.font = normal_font
    c.number_format = "0.0%"
    c.border = border
autosize(ws, 7, [16, 10, 14, 16, 14, 16, 14])

# =====================================================================
# SHEET 6: Failed Delivery Pareto
# =====================================================================
ws = wb.create_sheet("Failed Delivery Pareto")
r = title_block(ws, "Failed Delivery Root-Cause Pareto", "Cumulative % computed by formula (running total)")
for label, fname in [("By Region", "kpi_failed_delivery_region.csv"), ("By Category", "kpi_failed_delivery_category.csv")]:
    df = pd.read_csv(f"{DATA}/{fname}")
    key_col = df.columns[0]
    ws.cell(row=r, column=1, value=label).font = Font(name=FONT, bold=True, size=12, color=TEAL)
    r += 1
    headers = [key_col, "Failed Orders", "Reshipment Cost ($)", "Cumulative %"]
    for j, col in enumerate(headers, start=1):
        ws.cell(row=r, column=j, value=col)
    style_header(ws, r, 4)
    start_data = r + 1
    for i, (_, row_) in enumerate(df.iterrows()):
        rr = start_data + i
        ws.cell(row=rr, column=1, value=row_[key_col]).font = normal_font
        c = ws.cell(row=rr, column=2, value=int(row_["failed_orders"])); c.font = input_font; c.number_format = "#,##0"
        c = ws.cell(row=rr, column=3, value=round(float(row_["reshipment_cost"]), 2)); c.font = input_font; c.number_format = "$#,##0"
        if i == 0:
            c = ws.cell(row=rr, column=4, value=f"=B{rr}/SUM(B{start_data}:B{start_data+len(df)-1})")
        else:
            c = ws.cell(row=rr, column=4, value=f"=D{rr-1}+B{rr}/SUM(B{start_data}:B{start_data+len(df)-1})")
        c.font = normal_font
        c.number_format = "0.0%"
        for c_ in range(1, 5):
            ws.cell(row=rr, column=c_).border = border
            if i % 2 == 1:
                ws.cell(row=rr, column=c_).fill = PatternFill("solid", start_color=LIGHT, end_color=LIGHT)
    r = start_data + len(df) + 2
autosize(ws, 4, [22, 14, 18, 12])

# =====================================================================
# SHEET 7: Cross-Border USCA
# =====================================================================
ws = wb.create_sheet("Cross-Border USCA")
r = title_block(ws, "Cross-Border US <-> Canada (USCA Market)", "Post-customs on-time rate includes a modeled +1 day CBSA clearance buffer for Canada-bound orders")
df = pd.read_csv(f"{DATA}/kpi_cross_border_summary.csv")
df = df.rename(columns={"lane": "Lane", "orders": "Orders", "otif_rate": "OTIF Rate",
                         "on_time_rate_pre_customs_adj": "On-Time Rate (Pre-Customs)",
                         "on_time_rate_post_customs_adj": "On-Time Rate (Post-Customs)",
                         "cancel_rate": "Cancel Rate", "avg_cost_per_order": "Avg Cost/Order ($)", "avg_sales": "Avg Sales ($)"})
write_df(ws, df, start_row=r, pct_cols=["OTIF Rate", "On-Time Rate (Pre-Customs)", "On-Time Rate (Post-Customs)", "Cancel Rate"],
         money_cols=["Avg Cost/Order ($)", "Avg Sales ($)"], int_cols=["Orders"], widths=[32, 10, 12, 20, 20, 12, 16, 14])

# =====================================================================
# SHEET 8: Corporate Segment
# =====================================================================
ws = wb.create_sheet("Corporate Segment")
r = title_block(ws, "Corporate (B2B) Segment — Seasonal Volume & Category Mix", "Product-focus decision for this project: Corporate/B2B segment, per user direction")
ws.cell(row=r, column=1, value="Monthly Volume").font = Font(name=FONT, bold=True, size=12, color=TEAL)
r += 1
df = pd.read_csv(f"{DATA}/kpi_corporate_monthly.csv")
df = df.rename(columns={"order_month": "Month", "orders": "Orders", "total_units": "Total Units",
                         "total_sales": "Total Sales ($)", "otif_rate": "OTIF Rate", "avg_cost_per_order": "Avg Cost/Order ($)"})
r = write_df(ws, df, start_row=r, pct_cols=["OTIF Rate"], money_cols=["Total Sales ($)", "Avg Cost/Order ($)"],
             int_cols=["Orders", "Total Units"], widths=[12, 10, 12, 16, 12, 16])
r += 1
ws.cell(row=r, column=1, value="Category Mix").font = Font(name=FONT, bold=True, size=12, color=TEAL)
r += 1
df2 = pd.read_csv(f"{DATA}/kpi_corporate_category.csv")
df2 = df2.rename(columns={"Category Name": "Category", "orders": "Orders", "total_units": "Total Units",
                           "total_sales": "Total Sales ($)", "otif_rate": "OTIF Rate"})
write_df(ws, df2, start_row=r, pct_cols=["OTIF Rate"], money_cols=["Total Sales ($)"], int_cols=["Orders", "Total Units"],
         widths=[20, 10, 12, 16, 12])

# =====================================================================
# SHEET 9: Savings Simulation
# =====================================================================
ws = wb.create_sheet("Savings Simulation")
r = title_block(ws, "Cost Reduction Savings Simulation", "Three-lever plan: carrier diversification, zone skipping, packaging optimization (see report Section 12 / Appendix A.2)")
ws.cell(row=r, column=1, value="By Lever").font = Font(name=FONT, bold=True, size=12, color=TEAL)
r += 1
lv = pd.read_csv(f"{DATA}/kpi_savings_by_lever.csv")
lv = lv.rename(columns={"lever": "Lever", "annual_savings": "Annual Savings ($)", "pct_of_total_cost": "% of Total Cost"})
lv_start = r
r = write_df(ws, lv, start_row=r, money_cols=["Annual Savings ($)"], pct_cols=["% of Total Cost"], widths=[26, 18, 14])
# total row with formula
total_r = r
ws.cell(row=total_r, column=1, value="TOTAL").font = Font(name=FONT, bold=True)
data_first = lv_start + 1
data_last = lv_start + len(lv)
ws.cell(row=total_r, column=2, value=f"=SUM(B{data_first}:B{data_last})").font = Font(name=FONT, bold=True)
ws.cell(row=total_r, column=2).number_format = "$#,##0"
r = total_r + 2
ws.cell(row=r, column=1, value="By Region").font = Font(name=FONT, bold=True, size=12, color=TEAL)
r += 1
sr = pd.read_csv(f"{DATA}/kpi_savings_simulation.csv")
sr = sr.rename(columns={"Order Region": "Order Region", "current_total_cost": "Current Cost ($)",
                         "savings_carrier_diversification": "Savings: Carrier Diversification ($)",
                         "savings_zone_skipping": "Savings: Zone Skipping ($)",
                         "simulated_savings": "Total Simulated Savings ($)", "savings_pct": "Savings %"})
write_df(ws, sr, start_row=r, money_cols=["Current Cost ($)", "Savings: Carrier Diversification ($)",
         "Savings: Zone Skipping ($)", "Total Simulated Savings ($)"], pct_cols=["Savings %"],
         widths=[20, 16, 22, 18, 20, 12])

wb.save(OUT)
print("Workbook written to", OUT)
